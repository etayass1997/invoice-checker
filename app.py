import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(
    page_title="השוואת חשבוניות והזמנות",
    page_icon="📋",
    layout="centered"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;600;700&display=swap');
    * { direction: rtl; text-align: right; font-family: 'Heebo', sans-serif; }
    .title-block {
        background: linear-gradient(135deg, #1a3c6e 0%, #2c6fad 100%);
        padding: 2rem 2.5rem; border-radius: 12px; margin-bottom: 2rem; color: white;
    }
    .title-block h1 { color: white; font-size: 1.8rem; margin: 0; font-weight: 700; }
    .title-block p  { color: #cde; margin: 0.4rem 0 0; font-size: 0.95rem; }
    .upload-section {
        background: white; border-radius: 12px; padding: 1.5rem 2rem;
        margin-bottom: 1.2rem; box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        border-right: 4px solid #2c6fad;
    }
    .upload-section h3 { margin: 0 0 0.8rem; color: #1a3c6e; font-size: 1rem; }
    .settings-box {
        background: white; border-radius: 12px; padding: 1.5rem 2rem;
        margin-bottom: 1.2rem; box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        border-right: 4px solid #6c757d;
    }
    .summary-box {
        background: white; border-radius: 12px; padding: 1.5rem 2rem;
        margin: 1.2rem 0; box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    .summary-text {
        background: #f0f4f8; border-radius: 8px; padding: 1rem 1.2rem;
        font-size: 1rem; line-height: 1.8; color: #1a3c6e; margin-top: 1rem;
    }
    .stat-row { display: flex; gap: 1rem; justify-content: flex-end; flex-wrap: wrap; margin-top: 1rem; }
    .stat { background: #f0f4f8; border-radius: 8px; padding: 0.8rem 1.2rem; text-align: center; min-width: 110px; }
    .stat .num { font-size: 1.5rem; font-weight: 700; color: #1a3c6e; }
    .stat .lbl { font-size: 0.75rem; color: #666; margin-top: 2px; }
    div[data-testid="stButton"] button {
        background: linear-gradient(135deg, #1a3c6e, #2c6fad);
        color: white; border: none; border-radius: 8px; padding: 0.6rem 2rem;
        font-family: 'Heebo', sans-serif; font-size: 1rem; font-weight: 600;
        width: 100%; cursor: pointer;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="title-block">
    <h1>📋 השוואת חשבוניות והזמנות</h1>
    <p>השוואת סכומים בין דוח הזמנות לדוח חשבוניות</p>
</div>
""", unsafe_allow_html=True)

# ===== העלאת קבצים =====
col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="upload-section"><h3>📋 קובץ הזמנות</h3>', unsafe_allow_html=True)
    file_orders = st.file_uploader("הזמנות", type=['xlsx','xls'], key="orders", label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)
with col2:
    st.markdown('<div class="upload-section"><h3>🧾 קובץ חשבוניות</h3>', unsafe_allow_html=True)
    file_invoices = st.file_uploader("חשבוניות", type=['xlsx','xls'], key="invoices", label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

# ===== הגדרות =====
st.markdown('<div class="settings-box"><b>⚙️ הגדרות</b>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    vat_rate = st.number_input("שיעור מע\"מ (%)", min_value=0.0, max_value=30.0, value=18.0, step=0.1) / 100
with col2:
    diff_threshold = st.number_input("סף הפרש זניח (₪)", min_value=0.0, max_value=10.0, value=0.05, step=0.01)
st.markdown('</div>', unsafe_allow_html=True)


# ===== לוגיקת עיבוד =====
def validate_columns(df_ord, df_inv):
    required_orders = {
        'c_master_id':          'מזהה הזמנה ראשי',
        'c_reservation_id':     'מזהה חדר בהזמנה',
        'c_reservation_status': 'סטטוס הזמנה',
        'c_price_amount':       'סכום הזמנה',
    }
    required_invoices = {
        'c_folio_number': 'מספר פוליו',
        'invoice_amount': 'סכום חשבונית',
    }
    errors = []
    for col, desc in required_orders.items():
        if col not in df_ord.columns:
            errors.append(f"קובץ הזמנות — חסרה עמודה: **{col}** ({desc})")
    for col, desc in required_invoices.items():
        if col not in df_inv.columns:
            errors.append(f"קובץ חשבוניות — חסרה עמודה: **{col}** ({desc})")
    return errors

def process(file_orders, file_invoices, vat_rate, diff_threshold):
    df_ord_raw = pd.read_excel(file_orders)
    df_inv_raw = pd.read_excel(file_invoices)

    # בדיקת עמודות חובה
    errors = validate_columns(df_ord_raw, df_inv_raw)
    if errors:
        raise ValueError("שגיאה בקבצים:\n" + "\n".join(errors))

    # הזמנות checkout
    df_checkout = df_ord_raw[df_ord_raw['c_reservation_status'] == 'chkou'].copy()
    df_not_checkout = df_ord_raw[df_ord_raw['c_reservation_status'] != 'chkou'].copy()

    # מפתחות
    df_checkout['room_key'] = df_checkout['c_master_id'].astype(str) + df_checkout['c_reservation_id'].astype(str).str.zfill(2)
    df_inv_raw['room_key'] = df_inv_raw['c_folio_number'].astype(str).str[:9] + df_inv_raw['c_folio_number'].astype(str).str[-2:]

    inv_grouped = df_inv_raw.groupby('room_key').agg(
        סכום_חשבונית=('invoice_amount', 'sum'),
        ערוץ=('c_name', 'first'),
    ).reset_index()

    name_col = 'c_name' if 'c_name' in df_checkout.columns else None
    extra_cols = ['room_key','c_master_id','c_reservation_id','c_price_amount'] + ([name_col] if name_col else [])
    ord_clean = df_checkout[extra_cols].copy()
    new_cols = ['room_key','master_id','reservation_id','סכום_הזמנה'] + (['שם_אורח'] if name_col else [])
    ord_clean.columns = new_cols
    ord_clean['סכום_הזמנה_כולל_מעמ'] = (ord_clean['סכום_הזמנה'] * (1 + vat_rate)).round(2)

    # outer join
    merged = pd.merge(ord_clean, inv_grouped, on='room_key', how='outer')
    merged['הפרש'] = (merged['סכום_חשבונית'] - merged['סכום_הזמנה_כולל_מעמ']).round(2)

    # סטטוס עסקי
    def classify(row):
        no_order = pd.isna(row['master_id'])
        no_invoice = pd.isna(row['סכום_חשבונית'])
        diff = row['הפרש']
        if no_order:
            return 'חשבונית ללא הזמנה תואמת'
        if no_invoice:
            return 'הזמנה Checkout ללא חשבונית'
        if abs(diff) <= diff_threshold:
            return 'תקין'
        return 'הפרש סכומים — לבדיקה'

    merged['תוצאת בדיקה'] = merged.apply(classify, axis=1)

    # לשוניות
    valid       = merged[merged['תוצאת בדיקה'] == 'תקין']
    mismatches  = merged[merged['תוצאת בדיקה'] == 'הפרש סכומים — לבדיקה']
    no_invoice  = merged[merged['תוצאת בדיקה'] == 'הזמנה Checkout ללא חשבונית']
    no_order    = merged[merged['תוצאת בדיקה'] == 'חשבונית ללא הזמנה תואמת']

    # לא checkout
    nc_cols = ['c_master_id','c_reservation_id','c_reservation_status','c_price_amount'] + (['c_name'] if 'c_name' in df_not_checkout.columns else [])
    not_checkout_clean = df_not_checkout[nc_cols].copy()
    nc_names = ['master_id','reservation_id','סטטוס','סכום_הזמנה'] + (['שם_אורח'] if 'c_name' in df_not_checkout.columns else [])
    not_checkout_clean.columns = nc_names

    # סיכום
    checkout_count = len(df_checkout)
    summary = {
        'תאריך הפקת דוח': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'הזמנות Checkout': checkout_count,
        'תקינות': len(valid),
        'הפרשי סכומים': len(mismatches),
        'ללא חשבונית': len(no_invoice),
        'חשבונית ללא הזמנה': len(no_order),
        'לא Checkout': len(not_checkout_clean),
        'סך סכום הזמנות (כולל מעמ)': merged['סכום_הזמנה_כולל_מעמ'].sum().round(2),
        'סך סכום חשבוניות': merged['סכום_חשבונית'].sum().round(2),
        'סך הפרשים': merged['הפרש'].sum().round(2),
    }

    return summary, valid, mismatches, no_invoice, no_order, not_checkout_clean, merged


# ===== עיצוב גיליון =====
COLORS = {
    'header':   'FF1a3c6e',
    'תקין':     'FFd4edda',
    'הפרש':     'FFf8d7da',
    'חסר':      'FFfff3cd',
    'ללא':      'FFe1bee7',
    'neutral':  'FFf5f5f5',
    'summary':  'FFe8f0fe',
}

def style_sheet(ws, df, diff_col=None):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # כותרת
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor=COLORS['header'])
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # נתונים
    status_col_idx = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == 'תוצאת בדיקה':
            status_col_idx = i

    for row in ws.iter_rows(min_row=2):
        status = row[status_col_idx-1].value if status_col_idx else None
        if status == 'תקין':
            fill = COLORS['תקין']
        elif status == 'הפרש סכומים — לבדיקה':
            fill = COLORS['הפרש']
        elif 'ללא חשבונית' in str(status):
            fill = COLORS['חסר']
        elif 'ללא הזמנה' in str(status):
            fill = COLORS['ללא']
        else:
            fill = 'FFFFFFFF'

        for cell in row:
            cell.fill = PatternFill('solid', fgColor=fill)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border

        if diff_col and row[diff_col-1].value is not None:
            v = row[diff_col-1].value
            if isinstance(v, (int, float)) and abs(v) >= 0.05:
                row[diff_col-1].number_format = '+#,##0.00;-#,##0.00'

    # רוחב עמודות
    for i, cell in enumerate(ws[1], 1):
        col = get_column_letter(i)
        val = str(cell.value or '')
        ws.column_dimensions[col].width = max(14, min(30, len(val) * 1.8 + 4))


def build_excel(summary, valid, mismatches, no_invoice, no_order, not_checkout, merged):
    wb = Workbook()
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== גיליון 1: סיכום =====
    ws_sum = wb.active
    ws_sum.title = 'סיכום'

    ws_sum['A1'] = 'השוואת חשבוניות והזמנות'
    ws_sum['A1'].font = Font(bold=True, size=16, name='Arial', color='1a3c6e')
    ws_sum.merge_cells('A1:B1')
    ws_sum.row_dimensions[1].height = 30

    ws_sum['A3'] = 'מדד'
    ws_sum['B3'] = 'ערך'
    for cell in [ws_sum['A3'], ws_sum['B3']]:
        cell.fill = PatternFill('solid', fgColor=COLORS['header'])
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, (k, v) in enumerate(summary.items(), 4):
        ws_sum[f'A{i}'] = k
        ws_sum[f'B{i}'] = v
        for col in ['A', 'B']:
            ws_sum[f'{col}{i}'].fill = PatternFill('solid', fgColor=COLORS['summary'] if i % 2 == 0 else 'FFFFFFFF')
            ws_sum[f'{col}{i}'].font = Font(name='Arial', size=10)
            ws_sum[f'{col}{i}'].alignment = Alignment(horizontal='right')
            ws_sum[f'{col}{i}'].border = border

    # טקסט סיכום עסקי
    row_txt = len(summary) + 5
    txt = (f"מתוך {summary['הזמנות Checkout']} הזמנות Checkout נמצאו "
           f"{summary['תקינות']} התאמות תקינות, "
           f"{summary['הפרשי סכומים']} חריגות סכום, "
           f"{summary['ללא חשבונית']} הזמנות ללא חשבונית "
           f"ו-{summary['חשבונית ללא הזמנה']} חשבוניות ללא הזמנה תואמת.")
    ws_sum[f'A{row_txt}'] = txt
    ws_sum[f'A{row_txt}'].font = Font(name='Arial', size=11, bold=True, color='1a3c6e')
    ws_sum[f'A{row_txt}'].alignment = Alignment(wrap_text=True)
    ws_sum.merge_cells(f'A{row_txt}:B{row_txt}')
    ws_sum.row_dimensions[row_txt].height = 45

    ws_sum.column_dimensions['A'].width = 32
    ws_sum.column_dimensions['B'].width = 22

    # ===== שאר הגיליונות =====
    sheets = [
        ('תקינים',              valid,         'תוצאת בדיקה'),
        ('חריגים',              mismatches,    'תוצאת בדיקה'),
        ('חסר חשבונית',         no_invoice,    'תוצאת בדיקה'),
        ('חשבונית ללא הזמנה',   no_order,      'תוצאת בדיקה'),
        ('לא Checkout',       not_checkout,  None),
    ]

    cols_main = ['room_key','master_id','reservation_id','שם_אורח','ערוץ',
                 'סכום_הזמנה_כולל_מעמ','סכום_חשבונית','הפרש','תוצאת בדיקה']

    for title, df, status_col in sheets:
        ws = wb.create_sheet(title)
        if df.empty:
            ws['A1'] = 'אין רשומות'
            ws['A1'].font = Font(name='Arial', size=11, color='888888')
            continue

        # בחר עמודות רלוונטיות
        if status_col:
            show_cols = [c for c in cols_main if c in df.columns]
        else:
            show_cols = list(df.columns)

        data = df[show_cols].copy()

        # כתיבת כותרות
        for ci, col in enumerate(show_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = PatternFill('solid', fgColor=COLORS['header'])
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 28

        # כתיבת נתונים
        diff_col_idx = None
        for ci, col in enumerate(show_cols, 1):
            if col == 'הפרש':
                diff_col_idx = ci

        for ri, (_, row_data) in enumerate(data.iterrows(), 2):
            status_val = row_data.get('תוצאת בדיקה', '')
            if status_val == 'תקין':
                fill = COLORS['תקין']
            elif status_val == 'הפרש סכומים — לבדיקה':
                fill = COLORS['הפרש']
            elif 'ללא חשבונית' in str(status_val):
                fill = COLORS['חסר']
            elif 'ללא הזמנה' in str(status_val):
                fill = COLORS['ללא']
            else:
                fill = 'FFFFFFFF'

            for ci, col in enumerate(show_cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row_data[col])
                cell.fill = PatternFill('solid', fgColor=fill)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
                if diff_col_idx and ci == diff_col_idx:
                    v = row_data[col]
                    if isinstance(v, (int, float)) and abs(v) >= 0.05:
                        cell.number_format = '+#,##0.00;-#,##0.00'

        # שורת סיכום בתחתית
        total_row = len(data) + 2
        numeric_cols = {'סכום_הזמנה_כולל_מעמ', 'סכום_חשבונית', 'הפרש'}
        for ci, col in enumerate(show_cols, 1):
            cell = ws.cell(row=total_row, column=ci)
            if col in numeric_cols and col in data.columns:
                cell.value = data[col].sum().round(2)
                cell.number_format = '#,##0.00'
                if col == 'הפרש':
                    cell.number_format = '+#,##0.00;-#,##0.00'
            elif ci == 1:
                cell.value = 'סה"כ'
            cell.fill = PatternFill('solid', fgColor='FF1a3c6e')
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
        ws.row_dimensions[total_row].height = 24
        for ci, col in enumerate(show_cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(14, min(30, len(str(col)) * 1.8 + 4))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===== הרצה =====
if file_orders and file_invoices:
    if st.button("▶ הרץ השוואה"):
        with st.spinner("מעבד נתונים..."):
            try:
                summary, valid, mismatches, no_invoice, no_order, not_checkout, merged = process(
                    file_orders, file_invoices, vat_rate, diff_threshold
                )

                st.markdown(f"""
                <div class="summary-box">
                    <b>סיכום השוואה</b>
                    <div class="stat-row">
                        <div class="stat"><div class="num">{summary['הזמנות Checkout']}</div><div class="lbl">הזמנות Checkout</div></div>
                        <div class="stat"><div class="num" style="color:#28a745">{summary['תקינות']}</div><div class="lbl">תקינות ✅</div></div>
                        <div class="stat"><div class="num" style="color:#dc3545">{summary['הפרשי סכומים']}</div><div class="lbl">הפרשי סכומים 🔴</div></div>
                        <div class="stat"><div class="num" style="color:#ffc107">{summary['ללא חשבונית']}</div><div class="lbl">ללא חשבונית 🟡</div></div>
                        <div class="stat"><div class="num" style="color:#6f42c1">{summary['חשבונית ללא הזמנה']}</div><div class="lbl">חשבונית ללא הזמנה 🟣</div></div>
                    </div>
                    <div class="summary-text">
                        מתוך <b>{summary['הזמנות Checkout']}</b> הזמנות Checkout נמצאו
                        <b>{summary['תקינות']}</b> התאמות תקינות,
                        <b>{summary['הפרשי סכומים']}</b> חריגות סכום,
                        <b>{summary['ללא חשבונית']}</b> הזמנות ללא חשבונית
                        ו-<b>{summary['חשבונית ללא הזמנה']}</b> חשבוניות ללא הזמנה תואמת.
                    </div>
                </div>
                """, unsafe_allow_html=True)

                excel_data = build_excel(summary, valid, mismatches, no_invoice, no_order, not_checkout, merged)
                st.download_button(
                    label="📥 הורד דוח Excel מלא",
                    data=excel_data,
                    file_name=f"השוואת_חשבוניות_הזמנות_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                tab1, tab2, tab3, tab4 = st.tabs(["חריגים 🔴", "תקינים ✅", "חסר חשבונית 🟡", "חשבונית ללא הזמנה 🟣"])
                with tab1: st.dataframe(mismatches, use_container_width=True, hide_index=True)
                with tab2: st.dataframe(valid, use_container_width=True, hide_index=True)
                with tab3: st.dataframe(no_invoice, use_container_width=True, hide_index=True)
                with tab4: st.dataframe(no_order, use_container_width=True, hide_index=True)

            except ValueError as e:
                st.error("❌ שגיאה בקבצים")
                for line in str(e).split('\n'):
                    if line.strip():
                        st.warning(line)
            except Exception as e:
                st.error(f"❌ שגיאה בעיבוד הקבצים: {e}")
                st.exception(e)
else:
    st.info("יש להעלות את שני הקבצים כדי להתחיל")
